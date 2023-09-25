# read balance sheet and income statement data from a list of files
# copy select information to an output excel file.

# This code has been edited to ensure anonymity


import openpyxl

# function to read balance sheet data
def getBSData(inputFile,sheetName):

    # print('Getting balance sheet data')
    inputWb = openpyxl.load_workbook(inputFile, data_only = True)
    wSheet = inputWb[sheetName]

    bsData = [] # empty list for bs data

    # line and column paramters
    startRow = 14
    endRow = 149
    lineCol = 1
    itemCol = 2
    startAcctCol = 9
    endAcctCol = 10
    valuesCol = 23

    
    for r in range (startRow, endRow + 1):
        rowData = []
        rowData.append(wSheet.cell(row = r, column = lineCol).value)
        rowData.append(wSheet.cell(row = r, column = itemCol).value)
        rowData.append(wSheet.cell(row = r, column = startAcctCol).value)
        rowData.append(wSheet.cell(row = r, column = endAcctCol).value)
        rowData.append(wSheet.cell(row = r, column = valuesCol).value)

        bsData.append(rowData)

    inputWb.close()

    return bsData

# function to read income statement data
def getISData(inputFile,sheetName):

    # print('Getting income statement data')
    inputWb = openpyxl.load_workbook(inputFile, data_only = True)
    wSheet = inputWb[sheetName]

    isData = [] # empty list for is data

    # line and column paramters
    startRow = 6
    endRow = 449
    lineCol = 1
    itemCol = 2
    AcctCol = 9
    cpCol = 23
    ytdCol = 31

    
    for r in range (startRow, endRow + 1):
        rowData = []
        rowData.append(wSheet.cell(row = r, column = lineCol).value)
        rowData.append(wSheet.cell(row = r, column = itemCol).value)
        rowData.append(wSheet.cell(row = r, column = AcctCol).value)
        rowData.append(wSheet.cell(row = r, column = cpCol).value)
        rowData.append(wSheet.cell(row = r, column = ytdCol).value)

        isData.append(rowData)

    inputWb.close()

    return isData

# function to write data to output excel file
def write_output_file(outputFile, dataList):
    wb = openpyxl.Workbook() # create a blank workbook
    wbSheet = wb.active

    additionRow = 1
    for r in range (0, len(dataList)):
        columnCount = len(dataList[r])

        colNum = 1
        for c in range (0, columnCount):
            
            wbSheet.cell(row = additionRow, column = colNum).value = dataList[r][c]
            colNum += 1
            
        additionRow += 1

    wb.save(outputFile)
    wb.close()

# function to get get list of files to read from an excel file list#
# excel list contained statement type, inputfile name, tab name, and ouputfilename
def get_file_list():
    # print('getting file list')
    fileWb = openpyxl.load_workbook(r'filelist.xlsx')
    wSheet = fileWb['Sheet1']

    listData = []

    rowNum = 2
    endRow = wSheet.max_row + 1

    for rowNum in range (rowNum, endRow):
        rowData = []
        for colNum in range (1, 5):
                rowData.append(wSheet.cell(row = rowNum, column = colNum).value)
                colNum += 1
          
        listData.append(rowData)
        rowNum += 1

    fileWb.close()

    return listData
    

if __name__ == '__main__':


    listData = get_file_list()

    # cycle file list
    for i in range (0, len(listData)):
        statementType = listData[i][0]
        inputFile = listData[i][1]
        sheetName = listData[i][2]
        outputFile = listData[i][3]
        print(inputFile)

        # get balance sheet data
        if statementType == 'b':
            bsData = getBSData(inputFile,sheetName)
            write_output_file(outputFile, bsData)

        # get income statement data
        if statementType == 'i':
            isData = getISData(inputFile,sheetName)
            write_output_file(outputFile, isData)
    

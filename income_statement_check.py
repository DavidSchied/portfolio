# income_statement_check.py

# This program copies select income statement information from a file
# and copies it to another file with historical information.

# The purpose was to perform a check on the reasonableness of the 
# current period income statement

# This code has been edited to ensure anonymity

import openpyxl as xl

# dictionary of company ID's and company names
CompanyList = {1: 'Company1', 2: 'Company2', 3: 'Company3',
               4: 'Company4', 5: 'Company5', 6: 'Company6'}


# function gets a list of account names from the check file to then search the current period financials
def getAccountTitles(companyName):
    accountTitles = [] # empty list
    wb = xl.load_workbook(r'DirectoryName\Income_Statement_Check.xlsx')
    ws = wb[companyName] # select worksheet for the company you are working with
    # get account titles into a list
    for rowNum in range(1, ws.max_row + 1):
        cellValue = ws.cell(row = rowNum, column = 1).value
        if cellValue != None:
            cellValue = cellValue.lower().strip() # convert to lower case and remove whitespaces
            accountTitles.append(cellValue)

    return accountTitles
    
    wb.close()

# searches the current period financials for the list of accounts from the check file
def searchFinancials(fileName, accountTitles):
    wb = xl.load_workbook(fileName)
    ws = wb['Income Statement'] # select the income statement worksheet

    financialData = {}
    
    for rowNum in range(1, ws.max_row + 1):
        cellValue = ws.cell(row = rowNum, column = 1).value
        cellValue = str(cellValue)  # convert cellValue to a string
        cellValue = cellValue.lower().strip()   # convert to lower case and remove whitespaces
        # print(cellValue)

        for i in range(0, len(accountTitles)):
            if cellValue == accountTitles[i]:
                #print(cellValue)
                financialData[cellValue] = ws.cell(row = rowNum, column = 2).value

    return financialData

    wb.close()

# writes the dictionary of returned financial information to the check file
def writeDataBacktoCheckFile(companyID, results, period):
    wb = xl.load_workbook(r'DirectoryName\Income_Statement_Check.xlsx')
    ws = wb[companyID] # select worksheet for the company you are working with

    writeColumn = ws.max_column + 1
    
    ws.cell(row = 1, column = writeColumn).value = period # writes the month in the column header
    
    for rowNum in range(1, ws.max_row + 1):
        cellValue = ws.cell(row = rowNum, column = 1).value
        
        if cellValue != None:
            cellValue = cellValue.lower().strip()
            
            if cellValue in results:
                #print(results[cellValue])
                ws.cell(row = rowNum, column = writeColumn).value = results[cellValue]

    wb.save(r'DirectoryName\Income_Statement_Check.xlsx')
    wb.close()
    
if __name__ == '__main__':	
    print('Which period are you working with?')
    period = input('>')
    print()
    print('Which company are you working with?')
    inputCompany = int(input('> '))
    print(CompanyList[inputCompany])

    # fileLocation = input('Enter file location: ')
    print('Enter finanancial statements filename: ')
    fileName = input('> ')
    # fileName = fileLocation + '\\' + excelName

    # gets a list of account titles to search for from Income_Statement_Check.xlsx
    accountTitles = getAccountTitles(CompanyList[inputCompany])

    # get data from financial statements
    results = searchFinancials(fileName, accountTitles)
    
    # write results to check file
    writeDataBacktoCheckFile([inputCompany], results, period) 

    # end main program
